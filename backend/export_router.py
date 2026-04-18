"""
監査ログ エクスポート（Excel / PDF）
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db
from auth import get_current_user, UserContext
import io
from datetime import datetime

router = APIRouter(prefix="/export", tags=["export"])


def _fetch_logs(project_id: int, db: Session):
    rows = db.execute(
        text("""
            SELECT r.id, r.author, r.role, r.comment, r.status,
                   r.version, r.confirmed_at, r.confirmed_by,
                   to_char(r.created_at, 'YYYY-MM-DD HH24:MI') AS created_at,
                   p.name AS project_name
            FROM review_logs r
            JOIN projects p ON p.id = r.project_id
            WHERE r.project_id = :id
            ORDER BY r.created_at
        """),
        {"id": project_id}
    ).fetchall()
    return [dict(r._mapping) for r in rows]


@router.get("/{project_id}/excel")
def export_excel(
    project_id: int,
    db: Session = Depends(get_db),
    user: UserContext = Depends(get_current_user),
):
    user.require("export")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")

    logs = _fetch_logs(project_id, db)
    if not logs:
        raise HTTPException(status_code=404, detail="レビューログなし")

    wb = Workbook()
    ws = wb.active
    ws.title = "レビューラリー"

    # ── スタイル定義
    STATUS_FILL = {
        "issue":  PatternFill("solid", fgColor="FCE4EC"),
        "reply":  PatternFill("solid", fgColor="E8F5E9"),
        "review": PatternFill("solid", fgColor="FFF9C4"),
        "close":  PatternFill("solid", fgColor="F5F5F5"),
    }
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Meiryo")
    HEADER_FILL = PatternFill("solid", fgColor="1A2340")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "日時", "バージョン", "役割", "氏名", "種別", "コメント", "確定者", "確定日時"]
    col_widths = [5, 16, 10, 10, 14, 8, 60, 14, 18]

    # タイトル行
    project_name = logs[0]["project_name"]
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"レビューラリー監査ログ：{project_name}"
    title_cell.font = Font(bold=True, size=13, name="Meiryo")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"出力日時：{datetime.now().strftime('%Y-%m-%d %H:%M')}　件数：{len(logs)}件"
    ws["A2"].font = Font(size=9, color="666666", name="Meiryo")
    ws.row_dimensions[2].height = 16

    # ヘッダー行
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    STATUS_LABEL = {"issue": "指摘", "reply": "回答", "review": "再評価", "close": "確定"}

    for i, log in enumerate(logs, 1):
        row = i + 3
        values = [
            i,
            log["created_at"],
            f"v{log['version']}",
            log["role"],
            log["author"],
            STATUS_LABEL.get(log["status"], log["status"]),
            log["comment"],
            log["confirmed_by"] or "",
            log["confirmed_at"] or "",
        ]
        fill = STATUS_FILL.get(log["status"], PatternFill())
        aligns = ["center", "center", "center", "center", "center", "center", "left", "center", "center"]
        for col, (val, align) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name="Meiryo", size=9)
            cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=(col == 7))
        ws.row_dimensions[row].height = 40

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{len(logs)+3}"

    # 監査シート
    ws2 = wb.create_sheet("監査メタ")
    ws2["A1"] = "項目"
    ws2["B1"] = "値"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    meta = [
        ("プロジェクト", project_name),
        ("出力日時", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("総件数", len(logs)),
        ("指摘件数", sum(1 for l in logs if l["status"] == "issue")),
        ("回答件数", sum(1 for l in logs if l["status"] == "reply")),
        ("確定件数", sum(1 for l in logs if l["status"] == "close")),
    ]
    for r, (k, v) in enumerate(meta, 2):
        ws2[f"A{r}"] = k
        ws2[f"B{r}"] = v

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"review_log_pj{project_id}_{datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/{project_id}/pdf")
def export_pdf(
    project_id: int,
    db: Session = Depends(get_db),
    user: UserContext = Depends(get_current_user),
):
    user.require("export")
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable
        )
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    except ImportError:
        raise HTTPException(status_code=503, detail="reportlab not installed")

    logs = _fetch_logs(project_id, db)
    if not logs:
        raise HTTPException(status_code=404, detail="レビューログなし")

    # CIDフォント登録（日本語対応）
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    FONT = "HeiseiKakuGo-W5"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=20*mm, bottomMargin=20*mm
    )

    STATUS_LABEL = {"issue": "指摘", "reply": "回答", "review": "再評価", "close": "確定"}
    STATUS_COLOR = {
        "issue":  colors.HexColor("#FCE4EC"),
        "reply":  colors.HexColor("#E8F5E9"),
        "review": colors.HexColor("#FFF9C4"),
        "close":  colors.HexColor("#F5F5F5"),
    }

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", fontName=FONT, fontSize=14, spaceAfter=4)
    sub_style = ParagraphStyle("sub", fontName=FONT, fontSize=8, textColor=colors.grey)
    cell_style = ParagraphStyle("cell", fontName=FONT, fontSize=7.5, leading=11)

    project_name = logs[0]["project_name"]
    story = [
        Paragraph(f"レビューラリー監査ログ", title_style),
        Paragraph(f"対象PJ：{project_name}　出力日時：{datetime.now().strftime('%Y-%m-%d %H:%M')}　件数：{len(logs)}件", sub_style),
        Spacer(1, 4*mm),
        HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1A2340")),
        Spacer(1, 3*mm),
    ]

    # テーブルデータ
    header_row = ["#", "日時", "役割/氏名", "種別", "コメント"]
    table_data = [header_row]
    row_colors = [colors.HexColor("#1A2340")]

    for i, log in enumerate(logs, 1):
        comment_para = Paragraph(log["comment"], cell_style)
        role_text = log['role'] + chr(10) + log['author']
        role_para = Paragraph(role_text, cell_style)
        table_data.append([
            str(i),
            log["created_at"],
            role_para,
            STATUS_LABEL.get(log["status"], log["status"]),
            comment_para,
        ])
        row_colors.append(STATUS_COLOR.get(log["status"], colors.white))

    col_widths_pt = [10*mm, 28*mm, 24*mm, 14*mm, None]
    t = Table(table_data, colWidths=col_widths_pt, repeatRows=1)

    style_cmds = [
        ("FONTNAME",    (0, 0), (-1, -1), FONT),
        ("FONTSIZE",    (0, 0), (-1, -1), 7.5),
        ("FONTNAME",    (0, 0), (-1, 0),  FONT),
        ("FONTSIZE",    (0, 0), (-1, 0),  8),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1A2340")),
        ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
        ("ALIGN",       (0, 0), (3, -1),  "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [STATUS_COLOR.get(l["status"], colors.white) for l in logs]),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
    ]
    t.setStyle(TableStyle(style_cmds))
    story.append(t)

    doc.build(story)
    buf.seek(0)

    filename = f"review_log_pj{project_id}_{datetime.now().strftime('%Y%m%d%H%M')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )
